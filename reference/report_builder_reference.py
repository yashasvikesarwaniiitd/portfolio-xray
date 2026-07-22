"""
Portfolio X-Ray Report v3 — comprehension-first redesign
Layout system per section sheet:
  [banner] -> [what this section tells you] -> [data table | chart in its own zone]
  -> [HOW TO READ THIS box] -> [WHAT GOOD LOOKS LIKE line] -> [AI INSIGHT block]
No overlapping zones. One idea per chart. Every chart captioned.
"""
from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, PieChart, BarChart, LineChart, RadarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.fill import PatternFillProperties
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

NAVY="1F3A5F"; TEAL="2E8B8B"; GOLD="C9962A"; LIGHT="F4F6F8"; INK="22303C"
F_H1=Font(name="Calibri",bold=True,size=15,color="FFFFFF")
F_SUB=Font(name="Calibri",size=10,italic=True,color="FFFFFF")
F_H2=Font(name="Calibri",bold=True,size=12,color=NAVY)
F_KPI=Font(name="Calibri",bold=True,size=24,color=NAVY)
F_KPI_L=Font(name="Calibri",bold=True,size=10,color=TEAL)
F_BODY=Font(name="Calibri",size=10,color=INK)
F_MUTE=Font(name="Calibri",size=9,italic=True,color="7F8C8D")
F_NOTE=Font(name="Calibri",size=10,color=INK)
F_AI=Font(name="Calibri",size=10,color="26326B")
FILL_H=PatternFill("solid",fgColor=NAVY)
FILL_SUB=PatternFill("solid",fgColor=TEAL)
FILL_NOTE=PatternFill("solid",fgColor=LIGHT)
FILL_GOOD=PatternFill("solid",fgColor="EAF7EF")
FILL_AI=PatternFill("solid",fgColor="ECEFF9")
thin=Side(style="thin",color="D5DBE1")
BORD=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=Workbook()

def put_img(ws,name,anchor,scale=1.0):
    img=XLImage(f"/home/claude/charts/{name}.png")
    img.width=int(img.width*scale*0.62); img.height=int(img.height*scale*0.62)
    ws.add_image(img,anchor)

CHART_W=12.5; CHART_H=7.0   # cm — consistent everywhere
CHART_ROWS=16               # vertical rows a chart zone reserves

def sheet(title):
    ws=wb.active if wb.sheetnames==["Sheet"] else wb.create_sheet()
    ws.title=title; ws.sheet_view.showGridLines=False
    widths=[2,30,13,13,13,2,13,13,13,13,13,2]
    for i,w in enumerate(widths,start=1):
        ws.column_dimensions[get_column_letter(i)].width=w
    return ws

def banner(ws,title,what_it_tells):
    ws.merge_cells("A1:L2")
    c=ws["A1"]; c.value="  "+title; c.font=F_H1; c.fill=FILL_H
    c.alignment=Alignment(vertical="center")
    for r in (1,2):
        for col in range(1,13): ws.cell(row=r,column=col).fill=FILL_H
    ws.merge_cells("A3:L3")
    s=ws["A3"]; s.value="  WHAT THIS SECTION TELLS YOU:  "+what_it_tells
    s.font=F_SUB; s.fill=FILL_SUB
    for col in range(1,13): ws.cell(row=3,column=col).fill=FILL_SUB
    ws.row_dimensions[3].height=18
    return 5

def table(ws,r0,headers,rows,col0=2):
    for i,h in enumerate(headers):
        c=ws.cell(row=r0,column=col0+i,value=h)
        c.fill=FILL_SUB; c.font=Font(bold=True,color="FFFFFF",size=10); c.border=BORD
    for j,row in enumerate(rows,1):
        for i,v in enumerate(row):
            c=ws.cell(row=r0+j,column=col0+i,value=v); c.font=F_BODY; c.border=BORD
    return r0+len(rows)+1

def note_box(ws,r0,title,lines,fill,font,icon):
    ws.merge_cells(start_row=r0,start_column=2,end_row=r0,end_column=11)
    t=ws.cell(row=r0,column=2,value=f"{icon}  {title}")
    t.font=Font(bold=True,size=10,color=NAVY if fill!=FILL_AI else "26326B"); t.fill=fill
    for col in range(2,12): ws.cell(row=r0,column=col).fill=fill
    r=r0+1
    for ln in lines:
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=11)
        c=ws.cell(row=r,column=2,value="    "+ln); c.font=font; c.fill=fill
        c.alignment=Alignment(wrap_text=True,vertical="top")
        for col in range(2,12): ws.cell(row=r,column=col).fill=fill
        ws.row_dimensions[r].height=26
        r+=1
    return r+1

def how_to_read(ws,r0,lines): return note_box(ws,r0,"HOW TO READ THIS CHART",lines,FILL_NOTE,F_NOTE,"📖")
def good_looks(ws,r0,line):  return note_box(ws,r0,"WHAT GOOD LOOKS LIKE",[line],FILL_GOOD,F_NOTE,"✅")
def ai_block(ws,r0,lines,title="AI INSIGHT — written by the model after reading this section's numbers"):
    return note_box(ws,r0,title,["• "+l for l in lines],FILL_AI,F_AI,"🤖")

def style_bar(chart,color=TEAL):
    for s in chart.series:
        s.graphicalProperties.solidFill=color
        s.graphicalProperties.line.noFill=True

# ---------- sample data (real holding names, illustrative weights) ----------
ALLOC=[("Equity — India",44),("Equity — US",26),("MF — India",14),("Gold & Silver",7),("Crypto",6),("Other",3)]
GEO=[("India",58),("United States",32),("Global",10)]
CAP=[("Large cap",38),("Mid cap",21),("Small cap",29),("Micro / other",12)]
SECTORS=[("Technology",34),("Financials",17),("Commodities",12),("Industrials",11),("Consumer",9),("Energy",8),("Other",9)]
TOP=[("Nvidia",11.5),("Parag Parikh Flexi",8.2),("TCS",6.4),("IVV S&P 500",6.1),("Gold BeES",5.3),("QQQ",4.9),("Motilal Midcap",4.4),("Jio Financial",4.1)]
RISK_TIER=[("High",57),("Medium",33),("Low",10)]
BETAS=[("Bitcoin",2.6),("Nvidia",1.9),("Kotak Small Cap",1.3),("IVV S&P 500",1.0),("Parag Parikh",0.9),("TCS",0.7),("Gold BeES",0.1)]
OVER=[("NASDAQ 100 × FANG+",55),("IVV × QQQ",42),("Parag Parikh × Smallcap 250",9)]
COSTS=[("Axis Global FoF",1.24),("ICICI Manufacturing",0.98),("Parag Parikh Flexi",0.63),("Motilal Midcap",0.57),("Kotak Small Cap",0.44),("Tata Digital",0.31)]
RADAR=[("Equity",70,55),("Debt",0,20),("Gold",7,10),("International",32,15),("Cash/Alt",6,0)]
HHI=0.062; EFFN=16.1; N=43; SCORE=64

# ================= EXECUTIVE SUMMARY =================
ws=sheet("Executive Summary")
r=banner(ws,"Portfolio X-Ray — Executive Summary","Your whole portfolio in one page: four numbers, one score, and the AI's four most important findings.")
# KPI cards
kpis=[("Holdings",N,"things you own"),("Effective bets",EFFN,"what they behave like*"),
      ("Portfolio beta",1.28,"swing vs market"),("High-risk weight","57%","in your riskiest tier")]
col=2
for label,val,cap in kpis:
    ws.merge_cells(start_row=r,start_column=col,end_row=r,end_column=col+1)
    ws.cell(row=r,column=col,value=label).font=F_KPI_L
    ws.merge_cells(start_row=r+1,start_column=col,end_row=r+1,end_column=col+1)
    ws.cell(row=r+1,column=col,value=val).font=F_KPI
    ws.merge_cells(start_row=r+2,start_column=col,end_row=r+2,end_column=col+1)
    ws.cell(row=r+2,column=col,value=cap).font=F_MUTE
    for rr in range(r,r+3):
        for cc in range(col,col+2):
            ws.cell(row=rr,column=cc).border=BORD
    col+=2
    if col==6: col=7
ws.cell(row=r+3,column=2,value="*Effective bets = 1 / HHI. Overlapping funds and big positions make 43 holdings behave like ~16 independent decisions.").font=F_MUTE
ws.merge_cells(start_row=r+3,start_column=2,end_row=r+3,end_column=11)
# gauge
gr=r+5
put_img(ws,"gauge",f"B{gr}",1.0)

ws.cell(row=gr,column=5,value=f"Diversification score: {SCORE}/100").font=F_H2
ws.cell(row=gr+1,column=5,value="0–40 concentrated · 40–70 moderate · 70+ well spread").font=F_MUTE
r=gr+14
r=ai_block(ws,r,[
 "Three of your funds (QQQ, NASDAQ 100, FANG+) overlap up to 55% — they are effectively ONE bet wearing three names. This single fact explains most of the gap between 43 holdings and 16 effective bets.",
 "Your true Nvidia exposure is ~15%, not 11.5%: the direct stake plus Nvidia hiding inside three index funds. No single metric on any app shows you this — it needs look-through.",
 "You hold zero debt. Your only shock-absorber is a 7% gold sleeve against a 1.28-beta equity core — in a drawdown, almost everything you own falls together.",
 "One question controls more outcome than all others: is 32% international-tech a conviction, or an accident of buying similar funds on different apps?"],
 title="AI EXECUTIVE SUMMARY — the four findings that matter most, ranked by portfolio weight controlled")

# ================= HOW TO USE =================
ws=sheet("How To Use This Report")
r=banner(ws,"How To Use This Report","Two minutes here makes every other page self-explanatory.")
r=note_box(ws,r,"THE STRUCTURE OF EVERY SECTION",[
 "1. A teal strip states what the section tells you — read that first.",
 "2. A small table holds the raw numbers; the chart shows the same thing visually.",
 "3. 📖 HOW TO READ: exactly what the axes/slices mean and what to look for.",
 "4. ✅ WHAT GOOD LOOKS LIKE: a reference point so numbers have meaning.",
 "5. 🤖 AI INSIGHT: what the model noticed connecting THIS section to the others."],FILL_NOTE,F_NOTE,"🧭")
r=note_box(ws,r,"WHAT THIS REPORT WILL NEVER DO",[
 "It never says buy, sell, hold, or predicts prices. Every insight is a finding or a question.",
 "It stores nothing: your inputs are processed and discarded.",
 "Where data is estimated (overlap from top-10 holdings, beta from 2Y history) the Methodology page says so."],FILL_GOOD,F_NOTE,"🛡️")

# ================= ALLOCATION =================
ws=sheet("1 · Asset Allocation")
r=banner(ws,"1 · Asset Allocation","Where your money actually lives — by asset class, then geography, then company size.")
t=table(ws,r,["Asset class","Weight %"],ALLOC)
put_img(ws,"alloc",f"G{r}")

r=max(t,r+CHART_ROWS)
r=how_to_read(ws,r,[
 "Each slice is one asset class; the % label is its share of your portfolio.",
 "Look for: any single slice over half the ring, and whether any 'safety' slices (debt, gold) exist at all."])
r=good_looks(ws,r,"A resilient book usually has no class above ~60% and at least 15–25% in things that don't move with equities (debt, gold, cash).")
t=table(ws,r,["Geography","Weight %"],GEO)
put_img(ws,"geo",f"G{r}")

r=max(t,r+CHART_ROWS)
r=how_to_read(ws,r,["Slices are markets. Home-market share above ~70% = home bias; check whether it's a choice."])
t=table(ws,r,["Market cap","Weight %"],CAP)
put_img(ws,"cap",f"G{r}")

r=max(t,r+CHART_ROWS)
r=how_to_read(ws,r,["Taller bar = more money in that size of company. Small+micro together above ~30% marks an aggressive book."])
r=ai_block(ws,r,[
 "Small-cap 29% + micro 12% + crypto 6% puts ~47% of your book in high-volatility assets — nearly double what the crypto line alone suggests.",
 "Your India/US split looks diversified, but both sleeves lean tech — geography is masking a single global-technology factor bet.",
 "None of this is wrong; each tilt should simply be a decision you can state in one sentence. Can you, for each of the three?"])

# ================= DIVERSIFICATION =================
ws=sheet("2 · Diversification")
r=banner(ws,"2 · Diversification","How many INDEPENDENT bets you really own — the single most misunderstood number in investing.")
t=table(ws,r,["Dimension","You %","Balanced %"],RADAR)
put_img(ws,"radar",f"G{r}")

r=max(t,r+CHART_ROWS+2)
r=how_to_read(ws,r,[
 "The teal shape is YOUR allocation; the second shape is a balanced reference. Where your shape bulges past the reference = overweight; where it collapses inward = missing.",
 "One glance: your shape bulges on Equity & International and is flat on Debt — that flat edge IS the finding."])
r=good_looks(ws,r,"A balanced shape touches every axis. Shapes with a zero axis have a blind spot that shows up in drawdowns.")
t=table(ws,r,["Metric","Value","Plain English"],[
 ("Actual holdings",N,"lines in your portfolio"),
 ("HHI",HHI,"concentration index (lower = more spread)"),
 ("Effective holdings",EFFN,"how many independent bets they act like"),
 ("Positions under 0.5%",11,"too small to ever matter")])
r=ai_block(ws,r+1,[
 "43 holdings behaving like 16 is mostly ONE cause: the three overlapping US-tech funds. Fixing one overlap adds more real diversification than buying five new funds.",
 "11 positions are under 0.5% — even a double in any of them moves your portfolio by less than half a percent. They add tracking effort, not protection.",
 "Diversification of NAMES is not diversification of RISK: your radar's flat debt axis means every bulge falls together in an equity shock."])

# ================= CONCENTRATION =================
ws=sheet("3 · Concentration")
r=banner(ws,"3 · Concentration","Whether a handful of positions quietly control your outcome.")
cum=0; rows=[]
for n_,w_ in TOP:
    cum+=w_; rows.append((n_,w_,round(cum,1)))
t=table(ws,r,["Holding","Weight %","Cumulative %"],rows)
put_img(ws,"pareto",f"G{r}",1.05)

r=max(t,r+CHART_ROWS+1)
r=how_to_read(ws,r,[
 "Teal bars (left axis): each top holding's weight. Gold line (right axis): how they ADD UP as you go right.",
 "Look for: how quickly the gold line crosses 50%. Crossing it within 8 names = a top-heavy book."])
r=good_looks(ws,r,"Top-10 under ~50% and no single name above 10% is a common comfort zone. You: top-8 ≈ 51%, one name above 10%.")
r=ai_block(ws,r,[
 "The bars UNDERSTATE your real concentration: Nvidia's 11.5% direct stake ignores the Nvidia inside IVV, QQQ and FANG+ — look-through exposure is ~15%.",
 "This is why the HHI ('diversified, 0.062') and reality disagree: HHI cannot see inside funds. The report exists precisely for gaps like this.",
 "Question worth writing down: at what single-name exposure would you trim — 15%? 20%? If you can't name the line, you don't have one."])

# ================= SECTOR =================
ws=sheet("4 · Sector Exposure")
r=banner(ws,"4 · Sector Exposure","Which industries your outcome depends on — including the ones hiding inside funds.")
t=table(ws,r,["Sector","Weight %"],SECTORS)
put_img(ws,"sector",f"G{r}")

r=max(t,r+CHART_ROWS)
r=how_to_read(ws,r,["Longer bar = more of your money rides that industry. Compare the top bar to the rest — if it's ~2× the next, you have a dominant-sector book."])
r=good_looks(ws,r,"No sector above ~30% is the usual comfort line. You: Technology 34% stated — and closer to ~45% after counting tech-like financials.")
r=ai_block(ws,r,[
 "Jio Financial and Groww sit in 'Financials' but trade like technology — your economic tech exposure is ~45%, not 34%.",
 "Commodities + gold (~19% combined with the gold sleeve) is your only genuinely uncorrelated block; in a tech drawdown it is the only bar expected to hold.",
 "Test to apply: if AI capex disappoints for two quarters, which bars actually fall? If the answer is 'most of them', sector diversification is cosmetic."])

# ================= OVERLAP =================
ws=sheet("5 · Fund Overlap")
r=banner(ws,"5 · Fund Overlap","Different fund names can own the SAME stocks — you may hold fewer bets than you think.")
t=table(ws,r,["Fund pair","Overlap %"],OVER)
put_img(ws,"overlap",f"G{r}")

r=max(t,r+CHART_ROWS-2)
r=how_to_read(ws,r,["Each bar = a pair of your funds; length = % of holdings they share. Above ~40% means the pair is largely one bet with two fee lines."])
r=good_looks(ws,r,"Healthy pairs sit under ~20% overlap. One of your pairs is at 55%, one at 42% — both in the 'redundant' zone. Your Indian MF pair at 9% is genuinely doing diversification work.")
r=ai_block(ws,r,[
 "NASDAQ 100 × FANG+ at 55%: you pay two expense ratios for one exposure — the cheaper fund delivers the same bet with less drag.",
 "IVV × QQQ (42%) is structural — QQQ is a concentrated tilt of what IVV already holds. Fine if intentional; redundant if it just accumulated.",
 "Estimates use each fund's disclosed top-10 holdings (see Methodology) — real overlap is likely HIGHER, not lower."])

# ================= RISK =================
ws=sheet("6 · Risk Profile")
r=banner(ws,"6 · Risk Profile","How hard your portfolio swings when the market moves — and where the swing comes from.")
t=table(ws,r,["Holding","Beta vs NIFTY"],BETAS)
put_img(ws,"beta",f"G{r}")

r=max(t,r+CHART_ROWS)
r=how_to_read(ws,r,[
 "Beta 1.0 = moves with the market. 2.0 = swings twice as hard, both directions. 0.1 = barely moves with it.",
 "Look for: how much of your list sits above 1.0, and whether anything sits near 0 to cushion falls."])
r=good_looks(ws,r,"A balanced book mixes betas: growth above 1, ballast below 0.5. Your only sub-0.5 line is gold — a barbell with a thin safe end.")
r=ai_block(ws,r,[
 "Portfolio beta ~1.28: a NIFTY −15% quarter models to roughly −19% for you, before crypto's extra kick.",
 "57% of weight is in your own High tier — defensible at your horizon, but it should be a written decision, not a default.",
 "The most useful risk exercise this month: write your answer to 'at −19%, do I add, hold, or freeze?' BEFORE it happens. The answer defines whether this beta fits you."])

# ================= COST =================
ws=sheet("7 · Cost Drag")
r=banner(ws,"7 · Cost Drag","The only 'return' that is guaranteed — what your funds charge, compounded over a decade.")
t=table(ws,r,["Fund","Expense ratio %"],COSTS)
put_img(ws,"cost",f"G{r}")

r=max(t,r+CHART_ROWS)
r=how_to_read(ws,r,["Each bar = a fund's annual fee as % of your money in it. Index funds cluster near 0.1–0.3%; active funds must EARN the gap above that."])
yrs=[(f"Yr {i}", round(1000000*((1.0068)**i-1)/1000,1)) for i in (1,3,5,7,10)]
t=table(ws,r,["Year","Cumulative fees on ₹10L ('000)"],yrs)
put_img(ws,"feedrag",f"G{r}")

r=max(t,r+CHART_ROWS-2)
r=how_to_read(ws,r,["The line is money leaving quietly: cumulative fees on ₹10L at your weighted 0.68% rate. It curves upward because fees compound too."])
r=ai_block(ws,r,[
 "Axis Global FoF (1.24%) largely duplicates exposure you already own via IVV (~0.03%) — a ~40× fee for similar reach.",
 "Your weighted 0.68% is respectable, but it still compounds to ~₹70k per ₹10L over a decade — check the line chart's Yr-10 point.",
 "The one fee question that matters, per active fund: has it beaten its index twin AFTER fees over your holding period? Two of six currently haven't."])

# ================= QUESTIONS =================
ws=sheet("8 · Questions To Discuss")
r=banner(ws,"8 · Questions To Discuss","Not advice — the ranked agenda for your next conversation with yourself or an adviser.")
rows=[("HIGH","Is ~15% look-through Nvidia a conviction you would defend in writing?","Concentration"),
      ("HIGH","Do NASDAQ 100 + FANG+ + QQQ all need to exist, or is one doing the work of three?","Overlap"),
      ("MEDIUM","Is zero debt a decision about your horizon, or an app-default accident?","Diversification"),
      ("MEDIUM","Write the −19% answer: add, hold, or freeze?","Risk"),
      ("LOW","Which of the 11 sub-0.5% positions still earn a slot?","Diversification"),
      ("LOW","Is Axis Global FoF's 1.24% fee justified next to IVV?","Cost")]
t=table(ws,r,["Priority","Question","From section"],rows)
for i,(p,_,_) in enumerate(rows):
    ws.cell(row=r+1+i,column=2).fill=PatternFill("solid",fgColor={"HIGH":"FADBD8","MEDIUM":"FCF3CF","LOW":"D5F5E3"}[p])
ws.column_dimensions["C"].width=72
r=ai_block(ws,t+1,[
 "Ranked by portfolio weight controlled: the two HIGH questions govern ~30% of your money; the LOWs are hygiene.",
 "Every question here can legitimately be answered 'yes, intentional.' The report's job is to make sure each answer is CHOSEN, not defaulted."])

# ================= METHODOLOGY =================
ws=sheet("Methodology")
r=banner(ws,"Methodology & Boundaries","How every number is computed, where it's estimated, and the line this report never crosses.")
rows=[("Weights","User-provided or derived from amounts; nothing stored",""),
 ("HHI / Effective bets","Sum of squared weights; 1/HHI","Pure Python"),
 ("Beta","Daily-return regression vs NIFTY 50, ~2Y","yfinance"),
 ("Overlap","Shared names in disclosed TOP-10 holdings","Understates true overlap"),
 ("Expense ratios","Latest disclosed TER","AMC factsheets"),
 ("Look-through exposure","Direct weight + weight inside funds via top-10 disclosures","Estimate"),
 ("AI Insights","LLM reads each section's computed JSON; writes findings + questions","Refusal guardrail: any buy/sell verb regenerates the block")]
t=table(ws,r,["Item","Method","Note"],rows)
ws.cell(row=t+1,column=2,value="Hard boundary: never buy/sell/hold, never predictions. Educational analytics only. Not SEBI-registered advice.").font=Font(bold=True,size=10,color="C0392B")
ws.merge_cells(start_row=t+1,start_column=2,end_row=t+1,end_column=11)

out="/mnt/user-data/outputs/Portfolio_XRay_Report_v4.xlsx"
wb.save(out); print("saved",out,"| sheets:",len(wb.sheetnames))
