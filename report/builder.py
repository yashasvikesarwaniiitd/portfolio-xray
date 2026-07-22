"""Workbook builder — ported from reference/report_builder_reference.py, wired to real data.

Layout system per section sheet (kept exactly from the approved reference):
  [banner] -> [WHAT THIS SECTION TELLS YOU strip] -> [data table | chart zone]
  -> [📖 HOW TO READ box] -> [✅ WHAT GOOD LOOKS LIKE box] -> [🤖 AI INSIGHT block]

Tier gating: the SECTIONS registry maps each section to a min input level (L0/L1/L2).
Locked sections render as a sheet with an "add inputs to unlock" note. The level is
detected from the portfolio data present (report/data.py), never hardcoded.

generate_report() works even if the AI layer completely fails (deterministic fallback
insights) and even if fetches fail (sections degrade with honest notes, never crash).
"""
import os
import shutil
import tempfile
import time

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import agent
import logger
from report import charts, data, insights

NAVY = "1F3A5F"; TEAL = "2E8B8B"; LIGHT = "F4F6F8"; INK = "22303C"
F_H1 = Font(name="Calibri", bold=True, size=15, color="FFFFFF")
F_SUB = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
F_H2 = Font(name="Calibri", bold=True, size=12, color=NAVY)
F_KPI = Font(name="Calibri", bold=True, size=24, color=NAVY)
F_KPI_L = Font(name="Calibri", bold=True, size=10, color=TEAL)
F_BODY = Font(name="Calibri", size=10, color=INK)
F_MUTE = Font(name="Calibri", size=9, italic=True, color="7F8C8D")
F_NOTE = Font(name="Calibri", size=10, color=INK)
F_AI = Font(name="Calibri", size=10, color="26326B")
FILL_H = PatternFill("solid", fgColor=NAVY)
FILL_SUB = PatternFill("solid", fgColor=TEAL)
FILL_NOTE = PatternFill("solid", fgColor=LIGHT)
FILL_GOOD = PatternFill("solid", fgColor="EAF7EF")
FILL_AI = PatternFill("solid", fgColor="ECEFF9")
_thin = Side(style="thin", color="D5DBE1")
BORD = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

CHART_ROWS = 16  # vertical rows a chart zone reserves (reference constant)

# Registry: section key -> (sheet title, min input level). Rendered in this order.
SECTIONS = [
    ("allocation", "1 · Asset Allocation", "L1"),
    ("diversification", "2 · Diversification", "L1"),
    ("concentration", "3 · Concentration", "L1"),
    ("sector", "4 · Sector Exposure", "L1"),
    ("overlap", "5 · Fund Overlap", "L0"),
    ("risk", "6 · Risk Profile", "L2"),
    ("cost", "7 · Cost Drag", "L2"),
    ("questions", "8 · Questions To Discuss", "L1"),
]

TELLS = {
    "allocation": "Where your money actually lives — by asset class, then geography, then "
                  "company size.",
    "diversification": "How many INDEPENDENT bets you really own — the single most "
                       "misunderstood number in investing.",
    "concentration": "Whether a handful of positions quietly control your outcome.",
    "sector": "Which industries your outcome depends on — for the stocks you hold directly.",
    "overlap": "Different fund names can own the SAME stocks — you may hold fewer bets than "
               "you think.",
    "risk": "How hard your portfolio swings when the market moves — and where the swing "
            "comes from.",
    "cost": "The only 'return' that is guaranteed — what your funds charge, compounded over "
            "a decade.",
    "questions": "Not advice — the ranked agenda for your next conversation with yourself "
                 "or an adviser.",
}


# --- layout helpers (ported) -------------------------------------------------------------

def sheet(wb, title):
    ws = wb.active if wb.sheetnames == ["Sheet"] else wb.create_sheet()
    ws.title = title
    ws.sheet_view.showGridLines = False
    widths = [2, 30, 13, 13, 13, 2, 13, 13, 13, 13, 13, 2]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def banner(ws, title, what_it_tells):
    ws.merge_cells("A1:L2")
    c = ws["A1"]; c.value = "  " + title; c.font = F_H1; c.fill = FILL_H
    c.alignment = Alignment(vertical="center")
    for r in (1, 2):
        for col in range(1, 13):
            ws.cell(row=r, column=col).fill = FILL_H
    ws.merge_cells("A3:L3")
    s = ws["A3"]; s.value = "  WHAT THIS SECTION TELLS YOU:  " + what_it_tells
    s.font = F_SUB; s.fill = FILL_SUB
    for col in range(1, 13):
        ws.cell(row=3, column=col).fill = FILL_SUB
    ws.row_dimensions[3].height = 18
    return 5


def table(ws, r0, headers, rows, col0=2):
    for i, h in enumerate(headers):
        c = ws.cell(row=r0, column=col0 + i, value=h)
        c.fill = FILL_SUB; c.font = Font(bold=True, color="FFFFFF", size=10); c.border = BORD
    for j, row in enumerate(rows, 1):
        for i, v in enumerate(row):
            c = ws.cell(row=r0 + j, column=col0 + i, value=v)
            c.font = F_BODY; c.border = BORD
    return r0 + len(rows) + 1


def note_box(ws, r0, title, lines, fill, font, icon):
    ws.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=11)
    t = ws.cell(row=r0, column=2, value=f"{icon}  {title}")
    t.font = Font(bold=True, size=10, color=NAVY if fill != FILL_AI else "26326B")
    t.fill = fill
    for col in range(2, 12):
        ws.cell(row=r0, column=col).fill = fill
    r = r0 + 1
    for ln in lines:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
        c = ws.cell(row=r, column=2, value="    " + ln)
        c.font = font; c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(2, 12):
            ws.cell(row=r, column=col).fill = fill
        ws.row_dimensions[r].height = 26
        r += 1
    return r + 1


def how_to_read(ws, r0, lines):
    return note_box(ws, r0, "HOW TO READ THIS CHART", lines, FILL_NOTE, F_NOTE, "📖")


def good_looks(ws, r0, line):
    return note_box(ws, r0, "WHAT GOOD LOOKS LIKE", [line], FILL_GOOD, F_NOTE, "✅")


def ai_block(ws, r0, lines,
             title="AI INSIGHT — written by the model after reading this section's numbers"):
    return note_box(ws, r0, title, ["• " + l for l in lines], FILL_AI, F_AI, "🤖")


def put_img(ws, path, anchor, scale=1.0):
    if not path or not os.path.exists(path):
        return
    img = XLImage(path)
    img.width = int(img.width * scale * 0.62)
    img.height = int(img.height * scale * 0.62)
    ws.add_image(img, anchor)


def data_unavailable(ws, r, reason):
    return note_box(ws, r, "DATA UNAVAILABLE", [reason], FILL_NOTE, F_MUTE, "⚠️")


# --- section sheets ------------------------------------------------------------------------

def _sheet_allocation(ws, sec, ch, ai):
    r = banner(ws, "1 · Asset Allocation", TELLS["allocation"])
    t = table(ws, r, ["Asset class", "Weight %"], sec["asset_classes"])
    put_img(ws, ch.get("alloc"), f"G{r}")
    r = max(t, r + CHART_ROWS)
    r = how_to_read(ws, r, [
        "Each slice is one asset class; the % label is its share of your portfolio.",
        "Look for: any single slice over half the ring, and whether any 'safety' slices "
        "(debt, gold) exist at all."])
    r = good_looks(ws, r, "A resilient book usually has no class above ~60% and at least "
                          "15–25% in things that don't move with equities (debt, gold, cash).")
    t = table(ws, r, ["Geography", "Weight %"], sec["geography"])
    put_img(ws, ch.get("geo"), f"G{r}")
    r = max(t, r + CHART_ROWS)
    r = how_to_read(ws, r, ["Slices are markets. Home-market share above ~70% = home bias; "
                            "check whether it's a choice."])
    if sec.get("market_cap"):
        t = table(ws, r, ["Market cap", "Weight %"], sec["market_cap"])
        put_img(ws, ch.get("cap"), f"G{r}")
        r = max(t, r + CHART_ROWS)
        r = how_to_read(ws, r, ["Taller bar = more money in that size of company. Small+micro "
                                "together above ~30% marks an aggressive book."])
    ai_block(ws, r, ai)


def _sheet_diversification(ws, sec, ch, ai):
    r = banner(ws, "2 · Diversification", TELLS["diversification"])
    t = table(ws, r, ["Dimension", "You %", "Balanced %"], sec["radar"])
    put_img(ws, ch.get("radar"), f"G{r}")
    r = max(t, r + CHART_ROWS + 2)
    r = how_to_read(ws, r, [
        "The teal shape is YOUR allocation; the second shape is a balanced reference. Where "
        "your shape bulges past the reference = overweight; where it collapses inward = "
        "missing.",
        "One glance: a flat or missing axis IS the finding — that dimension is absent from "
        "your book."])
    r = good_looks(ws, r, "A balanced shape touches every axis. Shapes with a zero axis have "
                          "a blind spot that shows up in drawdowns.")
    t = table(ws, r, ["Metric", "Value", "Plain English"], [
        ("Actual holdings", sec["holdings_count"], "lines in your portfolio"),
        ("HHI", sec["hhi"], "concentration index (lower = more spread)"),
        ("Effective holdings", sec["effective_holdings"],
         "how many independent bets they act like"),
        ("Positions under 0.5%", sec["tiny_positions_under_half_pct"],
         "too small to ever matter"),
        ("Diversification score", f"{sec['diversification_score']}/100",
         "formula on the Methodology sheet")])
    put_img(ws, ch.get("gauge"), f"G{r}")
    ai_block(ws, max(t, r + 10) + 1, ai)


def _sheet_concentration(ws, sec, ch, ai):
    r = banner(ws, "3 · Concentration", TELLS["concentration"])
    rows = [(h["name"], h["weight_pct"], h["cumulative_pct"]) for h in sec["top_holdings"]]
    t = table(ws, r, ["Holding", "Weight %", "Cumulative %"], rows)
    put_img(ws, ch.get("pareto"), f"G{r}", 1.05)
    r = max(t, r + CHART_ROWS + 1)
    r = how_to_read(ws, r, [
        "Teal bars: each top holding's weight. Gold line: how they ADD UP as you go right.",
        "Look for: how quickly the gold line crosses 50%. Crossing it within 8 names = a "
        "top-heavy book."])
    n = len(rows)
    r = good_looks(ws, r, "Top-10 under ~50% and no single name above 10% is a common "
                          f"comfort zone. You: top-{n} ≈ {sec['top_n_cumulative_pct']:.0f}%, "
                          f"largest single name {sec['max_single_weight_pct']:.1f}%.")
    lt = sec.get("lookthrough")
    if lt:
        if lt.get("via"):
            lines = [f"{lt['stock']}: {lt['direct_pct']}% direct + {lt['via_funds_pct']}% "
                     f"inside your funds = ~{lt['total_pct']}% true exposure.",
                     "Via: " + "; ".join(f"{v['fund']} ({v['stock_pct_in_fund']}% of that "
                                         "fund)" for v in lt["via"])]
        else:
            lines = [f"{lt['stock']}: {lt['direct_pct']}% direct. "
                     + lt.get("note", "look-through estimate unavailable.")]
        r = note_box(ws, r, "LOOK-THROUGH EXPOSURE (direct + inside funds)", lines,
                     FILL_NOTE, F_NOTE, "🔍")
    ai_block(ws, r, ai)


def _sheet_sector(ws, sec, ch, ai):
    r = banner(ws, "4 · Sector Exposure", TELLS["sector"])
    pairs = list(sec["sectors_pct_of_direct_sleeve"].items())
    t = table(ws, r, ["Sector", "% of direct-stock sleeve"], pairs)
    put_img(ws, ch.get("sector"), f"G{r}")
    r = max(t, r + CHART_ROWS)
    r = how_to_read(ws, r, [
        "Longer bar = more of your direct-stock money rides that industry. Compare the top "
        "bar to the rest — if it's ~2× the next, you have a dominant-sector book.",
        f"Coverage: your direct-stock sleeve is {sec['direct_sleeve_pct_of_portfolio']}% of "
        "the portfolio; funds are mixed and not attributed here."])
    r = good_looks(ws, r, "No sector above ~30% of the sleeve is the usual comfort line.")
    ai_block(ws, r, ai)


def _sheet_overlap(ws, sec, ch, ai):
    r = banner(ws, "5 · Fund Overlap", TELLS["overlap"])
    if sec.get("pairs"):
        rows = [(p["pair"], p["overlap_pct"]) for p in sec["pairs"]]
        t = table(ws, r, ["Fund pair", "Overlap %"], rows)
        put_img(ws, ch.get("overlap"), f"G{r}")
        r = max(t, r + CHART_ROWS - 2)
    else:
        r = data_unavailable(ws, r, "No fund pair has disclosed top-10 holdings via free "
                                    "APIs — overlap is estimate-unavailable for this book.")
    r = how_to_read(ws, r, ["Each bar = a pair of your funds; length = % of holdings they "
                            "share. Above ~40% means the pair is largely one bet with two "
                            "fee lines."])
    r = good_looks(ws, r, "Healthy pairs sit under ~20% overlap. "
                          f"({sec.get('funds_with_disclosed_holdings', 0)} of "
                          f"{sec.get('funds_total', 0)} funds disclose top-10 holdings; "
                          "pairs without disclosure are not shown.)")
    r = note_box(ws, r, "ESTIMATE BASIS", [sec.get("note", "")], FILL_NOTE, F_MUTE, "ℹ️")
    ai_block(ws, r, ai)


def _sheet_risk(ws, sec, ch, ai):
    r = banner(ws, "6 · Risk Profile", TELLS["risk"])
    if sec.get("beta_ladder"):
        rows = [(b["name"], b["beta"]) for b in sec["beta_ladder"]]
        t = table(ws, r, ["Holding", "Beta vs NIFTY"], rows)
        put_img(ws, ch.get("beta"), f"G{r}")
        r = max(t, r + CHART_ROWS)
    else:
        r = data_unavailable(ws, r, "Beta could not be computed (price history unavailable).")
    r = how_to_read(ws, r, [
        "Beta 1.0 = moves with the market. 2.0 = swings twice as hard, both directions. "
        "0.1 = barely moves with it.",
        "Look for: how much of your list sits above 1.0, and whether anything sits near 0 "
        "to cushion falls."])
    r = good_looks(ws, r, "A balanced book mixes betas: growth above 1, ballast below 0.5.")
    t = table(ws, r, ["Risk tier (your labels)", "Weight %"], sec["risk_tiers_pct"])
    r = t + 1
    if sec.get("portfolio_beta_est"):
        r = note_box(ws, r, "PORTFOLIO BETA (estimated)", [
            f"~{sec['portfolio_beta_est']} weighted over the "
            f"{sec['beta_coverage_pct_of_portfolio']}% of the portfolio with beta data. A "
            f"NIFTY −15% quarter models to roughly "
            f"{sec['modelled_move_if_market_down_15pct']}% for that slice."],
            FILL_NOTE, F_NOTE, "📐")
    ai_block(ws, r, ai)


def _sheet_cost(ws, sec, ch, ai):
    r = banner(ws, "7 · Cost Drag", TELLS["cost"])
    if sec.get("expense_ratios"):
        rows = [(k["fund"], k["expense_ratio_pct"]) for k in sec["expense_ratios"]]
        t = table(ws, r, ["Fund", "Expense ratio %"], rows)
        put_img(ws, ch.get("cost"), f"G{r}")
        r = max(t, r + CHART_ROWS)
        r = how_to_read(ws, r, ["Each bar = a fund's annual fee as % of your money in it. "
                                "Index funds cluster near 0.1–0.3%; active funds must EARN "
                                "the gap above that."])
    else:
        r = data_unavailable(ws, r, "No fund in this book discloses its TER via free APIs.")
    if sec.get("ratio_unavailable_for"):
        r = note_box(ws, r, "TER NOT DISCLOSED (free APIs)",
                     [", ".join(sec["ratio_unavailable_for"])], FILL_NOTE, F_MUTE, "ℹ️")
    if sec.get("weighted_cost_pct_of_covered"):
        put_img(ws, ch.get("feedrag"), f"G{r}")
        r = how_to_read(ws, r + 1, [
            f"The line is money leaving quietly: cumulative fees on your "
            f"₹{sec['portfolio_value_for_feedrag']:,.0f} at the weighted "
            f"{sec['weighted_cost_pct_of_covered']}% rate (covered funds only). It curves "
            "upward because fees compound too."])
        r += CHART_ROWS - 4
    ai_block(ws, r, ai)


def _sheet_questions(ws, sec, ch, ai):
    r = banner(ws, "8 · Questions To Discuss", TELLS["questions"])
    qrows = [(q["priority"], q["question"], q["from_section"]) for q in sec["questions"]] \
        or [("—", "No flagged questions this run — the computed findings sit inside every "
                  "comfort line used in this report.", "—")]
    t = table(ws, r, ["Priority", "Question", "From section"], qrows)
    fills = {"HIGH": "FADBD8", "MEDIUM": "FCF3CF", "LOW": "D5F5E3"}
    for i, (p, _, _) in enumerate(qrows):
        if p in fills:
            ws.cell(row=r + 1 + i, column=2).fill = PatternFill("solid", fgColor=fills[p])
    ws.column_dimensions["C"].width = 72
    ai_block(ws, t + 1, ai)


_BUILDERS = {
    "allocation": _sheet_allocation, "diversification": _sheet_diversification,
    "concentration": _sheet_concentration, "sector": _sheet_sector,
    "overlap": _sheet_overlap, "risk": _sheet_risk, "cost": _sheet_cost,
    "questions": _sheet_questions,
}

_UNLOCK = {
    "L1": "Add invested amounts (a 'Total Invested' value per holding) to unlock this "
          "section — weights are needed to compute it.",
    "L2": "Add your monthly SIP amounts (the Mar'25…-style columns) to unlock this section "
          "— current values and exact units need the cashflow history.",
}


def _locked_sheet(ws, title, min_level, level):
    banner(ws, title, "This section is locked at your current input level.")
    note_box(ws, 5, f"LOCKED — needs input level {min_level} (you are at {level})",
             [_UNLOCK.get(min_level, "Add more portfolio inputs to unlock."),
              "Everything computed elsewhere in this report is unaffected."],
             FILL_NOTE, F_NOTE, "🔒")


# --- executive summary / how-to / methodology ---------------------------------------------

def _exec_summary(wb, bundle, ch, exec_bullets):
    ws = sheet(wb, "Executive Summary")
    r = banner(ws, "Portfolio X-Ray — Executive Summary",
               "Your whole portfolio in one page: four numbers, one score, and the AI's "
               "four most important findings.")
    div = bundle["sections"].get("diversification", {})
    risk = bundle["sections"].get("risk", {})
    kpis = [("Holdings", bundle["holdings_count"], "things you own"),
            ("Effective bets", div.get("effective_holdings", "n/a"),
             "what they behave like*"),
            ("Portfolio beta", risk.get("portfolio_beta_est", "n/a"), "swing vs market"),
            ("High-risk weight",
             f"{risk.get('high_risk_tier_pct', 'n/a')}%" if risk else "n/a",
             "in your riskiest tier")]
    col = 2
    for label, val, cap in kpis:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        ws.cell(row=r, column=col, value=label).font = F_KPI_L
        ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=col + 1)
        ws.cell(row=r + 1, column=col, value=val).font = F_KPI
        ws.merge_cells(start_row=r + 2, start_column=col, end_row=r + 2, end_column=col + 1)
        ws.cell(row=r + 2, column=col, value=cap).font = F_MUTE
        for rr in range(r, r + 3):
            for cc in range(col, col + 2):
                ws.cell(row=rr, column=cc).border = BORD
        col += 2
        if col == 6:
            col = 7
    ws.cell(row=r + 3, column=2,
            value="*Effective bets = 1 / HHI. Overlapping funds and big positions make "
                  f"{bundle['holdings_count']} holdings behave like "
                  f"~{div.get('effective_holdings', '?')} independent decisions.").font = F_MUTE
    ws.merge_cells(start_row=r + 3, start_column=2, end_row=r + 3, end_column=11)
    gr = r + 5
    put_img(ws, ch.get("gauge"), f"B{gr}", 1.0)
    if div.get("diversification_score") is not None:
        ws.cell(row=gr, column=5,
                value=f"Diversification score: {div['diversification_score']}/100").font = F_H2
        ws.cell(row=gr + 1, column=5,
                value="0–40 concentrated · 40–70 moderate · 70+ well spread").font = F_MUTE
    ai_block(ws, gr + 14, exec_bullets,
             title="AI EXECUTIVE SUMMARY — the four findings that matter most, ranked by "
                   "portfolio weight controlled")


def _how_to_use(wb, level):
    ws = sheet(wb, "How To Use This Report")
    r = banner(ws, "How To Use This Report",
               "Two minutes here makes every other page self-explanatory.")
    r = note_box(ws, r, "THE STRUCTURE OF EVERY SECTION", [
        "1. A teal strip states what the section tells you — read that first.",
        "2. A small table holds the raw numbers; the chart shows the same thing visually.",
        "3. 📖 HOW TO READ: exactly what the axes/slices mean and what to look for.",
        "4. ✅ WHAT GOOD LOOKS LIKE: a reference point so numbers have meaning.",
        "5. 🤖 AI INSIGHT: what the model noticed connecting THIS section to the others."],
        FILL_NOTE, F_NOTE, "🧭")
    r = note_box(ws, r, "WHAT THIS REPORT WILL NEVER DO", [
        "It never says buy, sell, hold, or predicts prices. Every insight is a finding or a "
        "question.",
        "It stores nothing: your inputs are processed and discarded.",
        "Where data is estimated (overlap from top-10 holdings, beta from 2Y history) the "
        "Methodology page says so."], FILL_GOOD, F_NOTE, "🛡️")
    note_box(ws, r, "YOUR INPUT LEVEL", [
        f"Detected level: {level}. L0 = holdings only · L1 = amounts invested (weights) · "
        "L2 = monthly SIP history (current values, risk & cost).",
        "Locked sections say exactly what input unlocks them."], FILL_NOTE, F_NOTE, "🔑")


def _methodology(wb, bundle):
    ws = sheet(wb, "Methodology")
    r = banner(ws, "Methodology & Boundaries",
               "How every number is computed, where it's estimated, and the line this "
               "report never crosses.")
    div = bundle["sections"].get("diversification", {})
    rows = [
        ("Input level", f"Detected {bundle['level']}: L1 = weights from amounts invested; "
                        "L2 = weights from reconstructed current values", "Never hardcoded"),
        ("Weights", "L2: exact unit reconstruction (units = SIP ÷ NAV-on-date) valued at "
                    "latest price; L1: invested-amount proportions", "Pure Python"),
        ("HHI / Effective bets", "Sum of squared weights; 1/HHI", "Pure Python"),
        ("Diversification score", div.get("score_formula",
         "40×min(effN/20,1) + 20×asset_spread + 20×geo_spread + 20×(1−overlap_penalty)"),
         "spread = 1−HHI over that grouping; overlap_penalty = max pairwise fund overlap"),
        ("Beta", "Daily-return regression vs NIFTY 50, ~2Y", "yfinance"),
        ("Overlap", "Shared names in disclosed TOP-10 holdings", "Understates true overlap"),
        ("Expense ratios", "Latest disclosed TER via yfinance; Indian MF TERs n/a on free "
                           "APIs", "Marked n/a where undisclosed"),
        ("Look-through exposure", "Direct weight + weight inside funds via top-10 "
                                  "disclosures", "Estimate"),
        ("Market-cap buckets", "INR: large ≥ ₹1L cr, mid ≥ ₹33k cr; USD: large ≥ $10B, "
                               "mid ≥ $2B", "Approximation"),
        ("AI Insights", "LLM reads each section's computed JSON; writes findings + "
                        "questions", "Guardrails: any buy/sell verb regenerates the block; "
                                     "every numeral must appear in the section JSON"),
    ]
    t = table(ws, r, ["Item", "Method", "Note"], rows)
    ws.cell(row=t + 1, column=2,
            value="Hard boundary: never buy/sell/hold, never predictions. Educational "
                  "analytics only. Not SEBI-registered advice.").font = \
        Font(bold=True, size=10, color="C0392B")
    ws.merge_cells(start_row=t + 1, start_column=2, end_row=t + 1, end_column=11)


# --- charts + insights orchestration -------------------------------------------------------

def _render_charts(bundle, outdir) -> dict:
    s = bundle["sections"]
    ch: dict = {}
    div = s.get("diversification", {})
    if div.get("diversification_score") is not None:
        ch["gauge"] = charts.render_gauge(div["diversification_score"], outdir)
        ch["radar"] = charts.render_radar(div.get("radar", []), outdir)
    alloc = s.get("allocation", {})
    if not alloc.get("error"):
        ch["alloc"] = charts.render_alloc(alloc.get("asset_classes", []), outdir)
        ch["geo"] = charts.render_geo(alloc.get("geography", []), outdir)
        ch["cap"] = charts.render_cap(alloc.get("market_cap", []), outdir)
    conc = s.get("concentration", {})
    if conc.get("top_holdings"):
        ch["pareto"] = charts.render_pareto(
            [(h["name"], h["weight_pct"]) for h in conc["top_holdings"]], outdir)
    sec = s.get("sector", {})
    if sec.get("sectors_pct_of_direct_sleeve"):
        ch["sector"] = charts.render_sector(
            list(sec["sectors_pct_of_direct_sleeve"].items()), outdir)
    ov = s.get("overlap", {})
    if ov.get("pairs"):
        ch["overlap"] = charts.render_overlap(
            [(p["pair"], p["overlap_pct"]) for p in ov["pairs"]], outdir)
    risk = s.get("risk", {})
    if risk.get("beta_ladder"):
        ch["beta"] = charts.render_beta(
            [(b["name"], b["beta"]) for b in risk["beta_ladder"]], outdir)
    cost = s.get("cost", {})
    if cost.get("expense_ratios"):
        ch["cost"] = charts.render_cost(
            [(k["fund"], k["expense_ratio_pct"]) for k in cost["expense_ratios"]], outdir)
    if cost.get("weighted_cost_pct_of_covered") and cost.get("portfolio_value_for_feedrag"):
        ch["feedrag"] = charts.render_feedrag(cost["weighted_cost_pct_of_covered"],
                                              cost["portfolio_value_for_feedrag"], outdir)
    return {k: v for k, v in ch.items() if v}


def _cross_context(sections: dict) -> dict:
    div = sections.get("diversification", {})
    conc = sections.get("concentration", {})
    return {
        "diversification_score": div.get("diversification_score"),
        "effective_holdings": div.get("effective_holdings"),
        "holdings_count": div.get("holdings_count"),
        "max_overlap_pct": sections.get("overlap", {}).get("max_overlap_pct"),
        "top_holding_weight_pct": conc.get("max_single_weight_pct"),
        "portfolio_beta_est": sections.get("risk", {}).get("portfolio_beta_est"),
        "high_risk_tier_pct": sections.get("risk", {}).get("high_risk_tier_pct"),
    }


def _build_insights(bundle, use_ai, usage) -> dict:
    sections = bundle["sections"]
    client = None
    if use_ai:
        try:
            import anthropic
            from dotenv import load_dotenv
            load_dotenv()
            client = anthropic.Anthropic()
        except Exception:
            client = None
    cross = _cross_context(sections)
    out: dict = {}
    for key, _, _ in SECTIONS:
        sec = sections.get(key)
        if not sec or sec.get("error"):
            continue
        if client:
            out[key] = insights.generate_insight(client, key, sec, cross, usage)
        else:
            out[key] = insights._fallback(key, sec)
    if client:
        out["__exec__"] = insights.generate_exec_summary(client, sections, usage)
    else:
        out["__exec__"] = insights._fallback("executive summary",
                                             {"sections_rendered": len(out)})
    return out


# --- entry point ---------------------------------------------------------------------------

def generate_report(portfolio_path: str = agent.DEFAULT_PORTFOLIO,
                    out_path: str = "Portfolio_Health_Report.xlsx",
                    use_ai: bool = True) -> dict:
    """Detect level -> assemble data -> render charts -> generate guarded insights -> build
    the workbook. Never blocked by the AI layer; sections degrade with honest notes."""
    started = time.perf_counter()
    bundle = data.assemble(portfolio_path)
    tmpdir = tempfile.mkdtemp(prefix="xray_charts_")
    try:
        ch = _render_charts(bundle, tmpdir)
        usage = logger.Usage()
        ins = _build_insights(bundle, use_ai, usage)

        wb = Workbook()
        _exec_summary(wb, bundle, ch, ins.get("__exec__", []))
        _how_to_use(wb, bundle["level"])
        locked = []
        for key, title, min_level in SECTIONS:
            ws = sheet(wb, title)
            sec = bundle["sections"].get(key)
            if data.LEVELS[bundle["level"]] < data.LEVELS[min_level]:
                _locked_sheet(ws, title, min_level, bundle["level"])
                locked.append(key)
            elif not sec or sec.get("error"):
                banner(ws, title, TELLS[key])
                data_unavailable(ws, 5, (sec or {}).get(
                    "error", "no data could be computed for this section"))
            else:
                _BUILDERS[key](ws, sec, ch, ins.get(key, []))
        _methodology(wb, bundle)
        wb.save(out_path)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)  # after save — openpyxl reads imgs at save

    cost = insights.log_report_cost(usage, started)
    return {"out_path": os.path.abspath(out_path), "level": bundle["level"],
            "sheets": len(SECTIONS) + 3, "locked_sections": locked,
            "ai_mode": "live" if use_ai and cost["input_tokens"] else "deterministic-fallback",
            "insight_cost": cost,
            "elapsed_s": round(time.perf_counter() - started, 1)}
